import calendar
import datetime
import time
import tkinter
import tkinter.messagebox
import sys

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Side
from openpyxl.utils import get_column_letter


# 结果Excel
class ResultExcel:

    def __init__(self, month=1):
        self._workbook = Workbook()
        self._month = month  # 需要几月份的工时统计
        self.employee_name_list = []  # 人员名单
        self.employee_department = ""  # 部门名称
        self.employee_num = 0  # 人员数量
        # 从配置文件.ini读入人员名单
        with open("配置文件.ini", 'r', encoding="utf8") as file_to_read:
            line = file_to_read.readline()  # 取出第一行，识别个人标识是否存在
            if line.find("研发中心-航电部-杨青云") == -1:
                tkinter.messagebox.showinfo("识别失败", "请勿删除或改动'配置文件.ini'的第一行")
                sys.exit()
            else:
                line = file_to_read.readline()  # 取出第二行，识别部门名称
                if line == '\n':
                    tkinter.messagebox.showinfo("未输入部门名", "请在第二行输入部门名称")
                    sys.exit()
                else:
                    self.employee_department = line.strip('\n')
                    line = file_to_read.readline()  # 剔除部门下面的一行
                    while True:
                        line = file_to_read.readline()
                        if not line:
                            break
                        line = line.strip('\n')
                        self.employee_name_list.append(line)
        self.employee_num = len(self.employee_name_list)  # 员工个数

    def init_template(self):
        month = self._month
        year = time.strftime("%Y")
        sheetname = str(month) + "月"  # 加上了-去除月份返回的前导0
        self._worksheet = self._workbook.active
        self._worksheet.title = sheetname
        header = ["序号", "姓名", "部门"]
        _, last_day_of_month = calendar.monthrange(int(year), int(month))  # 获取当前月份的最后一天
        month_day = list(range(1, last_day_of_month + 1, 1))
        month_day = [str(self._month) + "." + str(i) for i in month_day]
        header += month_day
        header.append("总计（分钟）")
        header.append("总计（小时）")
        for i in range(1, len(header) + 1, 1):  # 初始化表头
            self._worksheet.cell(1, i).value = header[i - 1]

        for i in range(1, self.employee_num + 1, 1):
            self._worksheet.cell(i + 1, 1).value = i
            self._worksheet.cell(i + 1, 2).value = self.employee_name_list[i - 1]
            self._worksheet.cell(i + 1, 3).value = self.employee_department

    def analyse(self, filename="全体员工.xlsx"):
        # 打开参考工作簿，注意只支持xlsx格式
        try:
            reference_wb = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            tkinter.messagebox.showerror("错误", filename + " 文件未在当前目录下找到")

        # 姓名从第2行，第2列开始
        for i in range(2, self._worksheet.max_row + 1):
            employee_name = self._worksheet.cell(i, 2).value  # 当前用户名
            for j in range(4, self._worksheet.max_column + 1):  # 从第2行，第4列开始填数据
                if j == self._worksheet.max_column - 1:  # 倒数第二列是总计（分钟）
                    col_letter = get_column_letter(j - 1)
                    self._worksheet.cell(i, j).value = "=SUM(D" + str(i) + ":" + col_letter + str(i) + ")"
                elif j == self._worksheet.max_column:  # 倒数第一列是总计（小时）
                    col_letter = get_column_letter(j - 1)
                    self._worksheet.cell(i, j).value = "=FLOOR(" + col_letter + str(i) + "/60,1)"
                else:
                    current_day = self._worksheet.cell(1, j).value  # 当前用户名下对应的日期
                    current_day = self.convert_date(current_day)  # 转换日期格式成参考表中的格式
                    reference_sheet = reference_wb.active
                    for ref_i in range(1, reference_sheet.max_row + 1):
                        # reference_sheet的第2列是日期，第6列是姓名，第11列是加班分钟，存在字符串格式的数字

                        if reference_sheet.cell(ref_i, 2).value == current_day and reference_sheet.cell(ref_i,
                                                                                                        6).value == employee_name and float(
                            reference_sheet.cell(ref_i, 11).value) != 0:
                            self._worksheet.cell(i, j).value = float(reference_sheet.cell(ref_i, 11).value)

    def convert_date(self, old_date):
        # old_date: 3.2
        # new_date: 2021-03-02
        temp = time.strftime("%Y") + "." + str(old_date)
        res = str(datetime.datetime.strptime(temp, "%Y.%m.%d"))
        return res[0:10]

    def beautify_excel(self):
        # 设置所有单元格数据居中，并绘制border
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in self._worksheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = align

        fille = openpyxl.styles.PatternFill('solid', fgColor='ffeb9c')
        # 标记所有为周末的本列为黄色（含周六和周日），日期从第四列开始，最后2列为总计列
        for col_j in range(4, self._worksheet.max_column - 1, 1):
            temp = self._worksheet.cell(1, col_j).value
            temp = temp.split('.')
            year = time.strftime("%Y")
            weekday = datetime.date(int(year), int(temp[0]), int(temp[1])).isoweekday()
            if weekday == 6 or weekday == 7:
                for row_i in range(1, self._worksheet.max_row + 1, 1):
                    self._worksheet.cell(row_i, col_j).fill = fille

    def save_excel(self):
        year = time.strftime("%Y")
        filename = str(year) + "年" + str(self._month) + "月份赶工核对（" + self.employee_department + "）.xlsx"
        self._workbook.save(filename)
        tkinter.messagebox.showinfo("成功", "已完成转换")


def main():
    def begin():
        resExcel = ResultExcel(month=Ent.get())
        resExcel.init_template()
        resExcel.analyse()
        resExcel.beautify_excel()
        resExcel.save_excel()

    rootGUI = tkinter.Tk()
    rootGUI.title("员工工时统计转换")
    L1 = tkinter.Label(rootGUI, text="请将【全体员工.xlsx】放入与本程序一个目录下\n编制者：【航电部-杨青云】，如有疑问请联系我！")
    L1.pack()
    L2 = tkinter.Label(rootGUI, text="清输入生成表月份")
    L2.pack()
    Ent = tkinter.Entry(rootGUI)
    Ent.pack()
    Btn = tkinter.Button(rootGUI, text="点我转换", command=begin)
    Btn.pack()

    rootGUI.mainloop()


# 按间距中的绿色按钮以运行脚本。
if __name__ == '__main__':
    main()
